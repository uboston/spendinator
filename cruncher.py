# Boston Underwood
# Clean up the downloaded csv that is in the folder by converting it to Pandas DataFrames and then writign to an excel file 
# with key data points. A spending by category by month bar chart, and other desired data visualizations.

# Import libraries
import pandas as pd
import os
import sys
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

# Gets the files in the directory
script_dir = os.path.dirname(os.path.abspath(__file__))
files_in_dir = os.listdir(script_dir)


# Filter only csv files
csv_files = [file for file in files_in_dir if file.lower().endswith('.csv')]
csvDict = {}
csvCounter = 1
csv_file_path = ""
bankAccount_path = ""
if csv_files:
    # Load up the dictionary
    for joma in csv_files:
        csvDict[csvCounter] = joma
        csvCounter += 1

    # Print the dictionary
    for key, value in csvDict.items():
        print(f"{key} : {value}")

    while True:
        try:
            creditCard = int(input("Please select which csv is from your credit card: "))
            if creditCard in csvDict:
                break
            else:
                print("\nWrong choice, try again ")
        except ValueError:
            print("\nPlease enter a number.")

    csv_file_path = os.path.join(script_dir, csvDict[creditCard])    
    del csvDict[creditCard]
    print("\n")

    if len(csvDict) > 1:
        for key, value in csvDict.items():
            print(f"{key} : {value}")
        
        while True:
            try:
                bankAccount = int(input("Please select which csv is from your checking account: "))
                if bankAccount in csvDict:
                    break
                else:
                    print("\nWrong choice, try again ")
            except ValueError:
                print("\nPlease enter a number.")
                    
        bankAccount_path = os.path.join(script_dir, csvDict[bankAccount])

    elif len(csvDict) > 0:
        bankAccount = next(iter(csvDict.keys()))
        bankAccount_path = os.path.join(script_dir, csvDict[bankAccount])

    else:
        print("Please add another csv ")
        sys.exit()
else:
    print("There are no csv files detected ")
    sys.exit()


try: 
    while True:
        try:
            hoursWorked = int(input("\nHow many hours are you working this pay period? "))
            if hoursWorked >= 0:
                break
            else: 
                print("\nEnter a positive number.")
        except ValueError:
            print("Please enter a number.")
except:
    print("Placeholder")


expectedIncome = 15.5 * hoursWorked
expectedIncome = expectedIncome * 0.9
healthandWellness = 20 / 2
personal = 15 / 2
gas = 160 / 2
rent = 650 / 2 
variableCost = (expectedIncome - healthandWellness - personal - gas - rent)

if variableCost < 0:
    variableCost = 0
    foodandDrink = 0
    entertainment = 0
    groceries  = 0
    shopping = 0
    extra = 0
else:
    foodandDrink = variableCost * 0.2
    entertainment = variableCost * 0.2
    groceries  = variableCost * 0.3
    shopping = variableCost * 0.2
    extra = variableCost * 0.1



wb = Workbook()
ws = wb.active
with open(f'{bankAccount_path}', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('test.xlsx')

dfBank = pd.read_excel('test.xlsx')
df = pd.read_csv(csv_file_path)
os.remove('test.xlsx')


# Clean up the file by deleting columns, rows, and fixing categories
# Delete columns, sort only purchases, and fix purchase categories

del df['Post Date']
del df['Memo']
df = df.query('Type == "Sale"')
del df['Type']
for transaction in df.index:
    if ((df.loc[transaction, "Category"] == "Gas") and (df.loc[transaction, "Amount"] > -15)):
        df.loc[transaction, "Category"] = "Food & Drink" 
for transaction in df.index:
    if df.loc[transaction, "Description"].startswith("Microsoft"):
        df.loc[transaction, "Category"] = "Entertainment"
for transaction in df.index:
    if (df.loc[transaction, "Description"].startswith(("BYU STORE"))) or (df.loc[transaction, "Description"].startswith(("VITALS"))):
        df.loc[transaction, "Category"] = "Education"

df['Amount'] = df['Amount'] * (-1)

budgetCategories = df['Category'].unique()
dfBudget = pd.DataFrame(budgetCategories, columns=['Budget'])
dfBudget.loc[dfBudget['Budget'] == 'Travel', 'Budget'] = 'Housing'
dfBudget.loc[dfBudget['Budget'] == 'Automotive', 'Budget'] = 'Misc'
dfBudget = dfBudget[~dfBudget['Budget'].isin(['Education'])]
dfBudget['Allotment'] = [f'{foodandDrink}', f'{entertainment}', f'{personal}', f'{groceries}', f'{gas}', f'{shopping}', f'{rent}', f'{healthandWellness}',f'{extra}']
dfBudget['Allotment'] = pd.to_numeric(dfBudget['Allotment'], errors='coerce')
dfBudget['Allotment'] = dfBudget['Allotment'].round(3)

# Cleans up the Bank Account DataFrame
del dfBank['Details']
del dfBank['Check or Slip #']
dfAccountExpenses = dfBank.copy()
del dfBank['Type']


# Creates a DataFrame according to the first entry of each month
dfprofit = dfBank.copy()
dfprofit['Posting Date'] = pd.to_datetime(dfprofit['Posting Date'])
dfprofit['Year'] = dfprofit['Posting Date'].dt.year
dfprofit['Month'] = dfprofit['Posting Date'].dt.month
dfprofit['Posting Date'] = dfprofit['Posting Date'].dt.date
# median_entries = dfprofit.groupby([dfprofit['Posting Date'].dt.year, dfprofit['Posting Date'].dt.month]).median().reset_index()
dfprofit = dfprofit.sort_values(by='Posting Date')
first_entries = dfprofit.groupby(['Year', 'Month']).first().reset_index()
first_entries["Posting Date"] = first_entries['Year'].astype(str) + '-' + first_entries['Month'].astype(str).str.zfill(2)
first_entries = first_entries.drop(columns=['Year', 'Month'])
first_entries['Profit'] = first_entries["Balance"] - first_entries['Balance'].shift(1)

# median_entries['Profit'] = median_entries["Balance"] - median_entries['Balance'].shift(1)
# print(median_entries)
# Create new DataFrames according to group sums by Month and Categories

# Create the Month DataFrame
df2 = df.copy()
df2['Transaction Date'] = pd.to_datetime(df2['Transaction Date'])
df2['YearMonth'] = df2['Transaction Date'].dt.to_period('M')
monthly_total = df2.groupby('YearMonth')['Amount'].sum().reset_index()
monthly_total.columns = ['Month', 'Total Amount']

# Create Category DateFrame
df3 = df.groupby('Category')['Amount'].sum().reset_index()

# Create Most Expensive Purchases for Each Category DataFrame
top3byCategory = df.groupby('Category').apply(lambda x: x.nlargest(3, 'Amount')).reset_index(drop=True)

dfCreditExpenses = df.copy()

del dfAccountExpenses ['Balance']
dfAccountExpenses = dfAccountExpenses.rename(columns={'Posting Date':'Date'})
dfBYU = dfAccountExpenses.copy()
dfBYU = dfBYU[dfBYU['Type'].isin(['ACH_CREDIT'])]
dfAccountExpenses = dfAccountExpenses[dfAccountExpenses['Type'].isin(['MISC_DEBIT', 'ACH_DEBIT', 'DEBIT_CARD'])]
dfAccountExpenses['Amount'] = dfAccountExpenses['Amount'] * (-1)
dfCreditExpenses = dfCreditExpenses.rename(columns={'Transaction Date':'Date'})
dfCreditExpenses = dfCreditExpenses[dfCreditExpenses['Category'] != 'Travel']
dfcombinedExpenses = pd.concat([dfCreditExpenses, dfAccountExpenses], axis=0)
dfcombinedExpenses['Date'] = pd.to_datetime(dfcombinedExpenses['Date'])
current_date = pd.Timestamp.now()
three_months_ago = current_date - pd.DateOffset(months=3)
dfcombinedExpenses = dfcombinedExpenses[dfcombinedExpenses['Date'] >= three_months_ago]
dfcombinedExpenses['Date'] = dfcombinedExpenses['Date'].dt.strftime('%m/%d/%Y')
dfBYU['Date'] = pd.to_datetime(dfBYU['Date'])
dfBYU = dfBYU[dfBYU['Date'] >= three_months_ago]
dfBYU['Date'] = dfBYU['Date'].dt.strftime('%m/%d/%Y')

# Add totals
sum_row = dfBYU[['Amount']].sum()
sum_row['Description'] = 'Total'
dfBYU = pd.concat([dfBYU, pd.DataFrame([sum_row])], ignore_index=True)

sum2_row = dfcombinedExpenses[['Amount']].sum()
sum2_row['Description'] = 'Total'
dfcombinedExpenses = pd.concat([dfcombinedExpenses, pd.DataFrame([sum2_row])], ignore_index=True)
# print(dfcombinedExpenses)


# Function to get the column letter based on numbers
def col_idx_to_excel_letter(col_idx):
    """Convert a 0-based column index to an Excel column letter."""
    letter = ''
    while col_idx >= 0:
        letter = chr(col_idx % 26 + ord('A')) + letter
        col_idx = col_idx // 26 - 1
    return letter

# Write the existing pd to an excel file
with pd.ExcelWriter('SpendingReport.xlsx', engine='xlsxwriter') as writer:
    df.to_excel(writer, sheet_name='Analysis', index=False, startcol= 0)
    monthly_total.to_excel(writer, sheet_name="Analysis", index=False, startcol= len(df.columns)+ 2)
    df3.to_excel(writer, sheet_name="Analysis", index=False, startcol= len(df.columns) +  6)
    top3byCategory.to_excel(writer, sheet_name="Analysis", index=False, startcol=len(df.columns) + 10)
    dfBank.to_excel(writer, sheet_name="Bank Balance", index=False, startcol=0)
    first_entries.to_excel(writer, sheet_name="Bank Balance", index=False, startcol=len(dfBank.columns) + 2)
    dfcombinedExpenses.to_excel(writer, sheet_name="Analysis", index=False, startcol=len(df.columns) + 15)
    dfBYU.to_excel(writer, sheet_name="Bank Balance", index=False, startcol=len(dfBank.columns) + 8)
    dfBudget.to_excel(writer, sheet_name="Analysis", index=False, startcol= len(df.columns) + 20)

    workbook = writer.book
    worksheet = writer.sheets['Analysis']
    worksheet2 = writer.sheets['Bank Balance']
    worksheet3 = workbook.add_worksheet('Dashboard')

    currency = workbook.add_format({'num_format': '$#,##0.00'})

    worksheet.set_column('D:D', None, currency)
    worksheet.set_column('H:H', None, currency)
    worksheet.set_column('L:L', None, currency)
    worksheet.set_column('R:R', None, currency)
    worksheet.set_column('W:W', None, currency)
    worksheet.set_column('Z:Z', None, currency)
    worksheet2.set_column('C:D', None, currency)
    worksheet2.set_column('O:O', None, currency)
    worksheet2.set_column('I:K', None, currency)
    

    categoryChart = workbook.add_chart({'type': 'column'})
    chart2 = workbook.add_chart({'type': 'column'})
    chart = workbook.add_chart({'type': 'column'})
    netChart = workbook.add_chart({'type': 'column'})
    balanceChart = workbook.add_chart({'type': 'line'})
    profitChart = workbook.add_chart({'type': 'bar'})
    budgetChart = workbook.add_chart({'type': 'pie'})

    # print(dir(workbook))
    months_col = col_idx_to_excel_letter(len(df.columns) + 2)
    value_col = col_idx_to_excel_letter(len(df.columns) + 3)
    chart_position ='A1'

    category_col = col_idx_to_excel_letter(len(df.columns) + 6)
    sums_col = col_idx_to_excel_letter(len(df.columns) + 7)
    chart_position2 = 'H1'

    # print(f'Category column: {category_col}')
    # print(f'Value column: {sums_col}')
    # print(f'Chart position 1: {chart_position}')
    # print(f'Chart position 2: {chart_position2}')
    # print(f'{category_col}2:{category_col}{len(df3) + 1}')
    # print(f'{sums_col}2:{sums_col}{len(df3) + 1}')
    
    
    chart.add_series({
    'name': 'Amount Spent',
    'line': {'color': 'blue'},
    'categories': f'=Analysis!${months_col}$2:${months_col}${len(monthly_total) + 1}',
    'values': f'=Analysis!${value_col}$2:${value_col}${len(monthly_total) + 1}'})


    chart.set_title({'name': 'Monthly Spending'})
    chart.set_x_axis({'name': 'Month'})
    chart.set_y_axis({'name': '$$$ Spent'})

    chart2.add_series({
        'name': 'Amount Spent',
        'categories': f'=Analysis!${category_col}$2:${category_col}${len(df3) + 1}',
        'values': f'=Analysis!${sums_col}$2:${sums_col}${len(df3) + 1}'
    })

    chart2.set_title({'name': 'Spending by Category'})
    chart2.set_x_axis({'name': 'Category'})
    chart2.set_y_axis({'name': '$$$ Spent'})

    descriptcolumnLetter = col_idx_to_excel_letter(len(df.columns) + 11)
    amountcolumnLetter = col_idx_to_excel_letter(len(df.columns) + 13)

    kategories = top3byCategory['Category'].unique()
    for category in (kategories):
        category_df = top3byCategory[top3byCategory['Category'] == category]
        start_row = category_df.index[0] + 2
        end_row = start_row + len(category_df) - 1
        
        categoryChart.add_series({
            'name': category,
            'categories': f'=Analysis!${descriptcolumnLetter}${start_row}:${descriptcolumnLetter}${end_row}',
            'values': f'=Analysis!${amountcolumnLetter}${start_row}:${amountcolumnLetter}${end_row}',
        })


    chart_position3 = "A16"
    categoryChart.set_title({'name': 'Top 3 Purchases by Category'})
    categoryChart.set_x_axis({'name': 'Category'})
    categoryChart.set_y_axis({'name': '$$$ Spent'})

    profitDesc_col = col_idx_to_excel_letter(len(dfBank.columns) + 3)
    profitValue_col = col_idx_to_excel_letter(len(dfBank.columns) + 6)

    netChart.add_series({
        'name': 'NET +-',
        'values': f'=Bank Balance!${profitValue_col}2:${profitValue_col}{len(first_entries) + 1}',
    })

    chart_position4 = "H16"
    netChart.set_title({'name': 'NET $$'})
    netChart.set_x_axis({'name': 'Month', 'label_position': 'none', 'visible': False})
    netChart.set_y_axis({'name': 'Change in Balance'})

    balanceChart.add_series({
        'name': 'Balance',
        'values': f'=Bank Balance!$D2:$D{len(dfBank) + 1}',
    })

    chart_position5 = "O1"
    balanceChart.set_title({'name': 'Balance Over Time'})
    balanceChart.set_x_axis({'name': 'Time', 'visible': False, 'reverse': True})
    balanceChart.set_y_axis({'name': 'Balance $$'})

    byuCol = col_idx_to_excel_letter(len(dfBank.columns) + 10)

    profitChart.add_series({
        'name': 'BYU Income',
        'values': f'=Bank Balance!${byuCol}{len(dfBYU) + 1}',
    })

    expenseSumCol = col_idx_to_excel_letter(22)
    
    profitChart.add_series({
        'name': 'Total Expenses',
        'values': f'=Analysis!${expenseSumCol}{len(dfcombinedExpenses) + 1}',
    })

    chart_position6 = "O16"
    profitChart.set_title({'name': 'Expenses v.  BYU Income'})
    profitChart.set_x_axis({'name': 'Last 3 Months'})
    profitChart.set_y_axis({'name': 'Amount $'})


    budCol = col_idx_to_excel_letter(len(df.columns) + 20)
    alotCol = col_idx_to_excel_letter(len(df.columns) + 21)

    budgetChart.add_series({
        'name': 'Budget',
        'categories': f'=Analysis!${budCol}2:${budCol}{len(dfBudget) + 1}',
        'values': f'=Analysis!${alotCol}2:${alotCol}{len(dfBudget) + 1}',
        'data_labels': {
            'value': True,
            'category': True,
            'leader_lines': True,
            'position': 'best_fit',
            },
    })


    chart_position7 = "V1"
    budgetChart.set_title({'name': 'Budget'})
    budgetChart.set_x_axis({'name': 'Categories'})
    budgetChart.set_y_axis({'name': 'Amount $'})


    worksheet3.insert_chart(chart_position, chart)
    worksheet3.insert_chart(chart_position2, chart2)
    worksheet3.insert_chart(chart_position3, categoryChart)
    worksheet3.insert_chart(chart_position4, netChart)
    worksheet3.insert_chart(chart_position5, balanceChart)
    worksheet3.insert_chart(chart_position6, profitChart)
    worksheet3.insert_chart(chart_position7, budgetChart)

wb = load_workbook('SpendingReport.xlsx')
wb.active = wb['Dashboard']
wb.save('SpendingReport.xlsx')

# workbook.close()
os.system("open SpendingReport.xlsx")